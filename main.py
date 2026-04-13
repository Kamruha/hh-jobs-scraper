from bs4 import BeautifulSoup
import requests
import time
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# --- Request headers to mimic a real browser ---
headers = {
    'User-Agent': 'Mozilla/5.0'
}

# --- Storage for all collected vacancies ---
data = []

# --- Set to track seen vacancies and avoid duplicates ---
seen = set()

# --- Main scraping loop: iterate through pages ---
for page in range(1, 20):
    print(f'Parsing page {page} of 20...')

    # Build URL with current page number, search by vacancy name only
    url = f'https://hh.ru/search/vacancy?text=python+developer&search_field=name&page={page}'

    # Send request with timeout; skip page on network error or bad status
    try:
        response = requests.get(url, headers=headers, timeout=10)
        if response.status_code != 200:
            print(f'Page {page}: status {response.status_code}, skipping...')
            continue
    except requests.exceptions.RequestException as e:
        print(f'Request error on page {page}: {e}')
        continue

    # Parse HTML response
    soup = BeautifulSoup(response.text, 'html.parser')

    # Find all vacancy cards on the page
    container = soup.find_all('div', {'data-qa': 'vacancy-serp__vacancy'})

    # If no cards found — we've reached the last page, stop
    if not container:
        print(f'No more vacancies at page {page}, stopping...')
        break

    # --- Loop through each vacancy card ---
    for con in container:

        # Extract individual fields from the card
        title = con.find('a', {'data-qa': 'serp-item__title'})
        company = con.find('span', {'data-qa': 'vacancy-serp__vacancy-employer-text'})
        location = con.find('span', {'data-qa': 'vacancy-serp__vacancy-address'})
        salary = con.find('span', class_='magritte-text___pbpft_5-0-1 magritte-text_style-primary___AQ7MW_5-0-1 magritte-text_typography-label-1-regular___pi3R-_5-0-1')

        # Get text values, fall back to placeholder if element not found
        title_tag = title.text if title else 'no title found'
        company_tag = company.text if company else 'no company found'
        location_tag = location.text if location else 'no location found'

        # Get raw salary text, then strip everything except digits and currency symbols
        salary_tag = salary.text if salary else 'no salary found'
        salary_tag = re.sub(r'[^\d\s\-–₽$€]', '', salary_tag).strip() if salary else 'no salary found'

        # Extract direct link to the vacancy from the title element
        link_tag = title['href'] if title else 'no link found'

        # Skip duplicate vacancies (same title + company)
        key = (title_tag, company_tag)
        if key in seen:
            continue
        seen.add(key)

        # Append vacancy data to the list
        data.append({
            'Title': title_tag,
            'Company': company_tag,
            'Location': location_tag,
            'Salary': salary_tag,
            'Link': link_tag
        })

    # Polite delay between requests to avoid getting blocked
    time.sleep(1)

# --- Save collected data to Excel ---
df = pd.DataFrame(data)
df.to_excel('vacancies.xlsx', index=False)

# --- Auto-fit column widths for readability ---
wb = load_workbook('vacancies.xlsx')
ws = wb.active

for column in ws.columns:
    max_length = max(len(str(cell.value or '')) for cell in column)
    ws.column_dimensions[column[0].column_letter].width = max_length + 4

# --- Make links in the Link column clickable ---
for cell in ws['E']:
    if cell.value and cell.value.startswith('http'):
        cell.hyperlink = cell.value
        cell.font = Font(color='0000FF', underline='single')

wb.save('vacancies.xlsx')

print(f'\nDone! {len(data)} vacancies saved to vacancies.xlsx')
